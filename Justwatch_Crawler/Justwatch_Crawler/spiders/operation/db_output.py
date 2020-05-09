import MySQLdb
import sys
import os
import db_detail
import datetime
from openpyxl import Workbook
from datetime import datetime,timedelta

class db_output_stats:

    def __init__(self):  
        self.db_list=[]
        self.connection=''
        self.cursor=''
        self.movie_fieldnames=["Movie_id","Title","Show_type","Description","Release_year","original_title","OTT","Cast","Duration","Genres","Rating","Age_Rating","Service_name","Added_to_site","Updated_at"]
        self.series_fieldnames=["Series_id","Season_id","Title","Show_type","Description","Release_year","original_title","Cast","Season_number","Genres","Rating","Age_Rating","Service_name","Added_to_site","Updated_at"]
        self.episode_fileldnames=["Series_id","Season_id","Episode_id","Series_title","Title","Show_type","Description ","OTT","Duration","season_number","Episode_number","Service_name","Updated_at"]

    def set_up_db_connection(self):
        self.connection=MySQLdb.connect(host=db_detail.IP_addr,user="%s"%db_detail.username,passwd="%s"%db_detail.passwd,db=db_detail.database_name)
        self.cursor=self.connection.cursor()     

    def main(self):
        self.set_up_db_connection()
        result_sheet='/operation/attachments/Justwatch_keepup_contents_%s.xlsx'%(datetime.now().strftime('%b %d, %Y'))
        excel_file = Workbook()
        if (os.path.isfile(os.getcwd()+result_sheet)):
            os.remove(os.getcwd()+result_sheet)

        excel_sheet = excel_file.create_sheet(title='Movies', index=0)
        excel_sheet.append(self.movie_fieldnames)
        query="select Movie_id,Title,Show_type,Description,Release_year,original_title,OTT,Cast,Duration,Genres,Rating,Age_Rating,Service_name,Added_to_site,Updated_at from {table_name} where Updated_at='%s'".format(table_name=db_detail.movie_table)
        self.cursor.execute(query%(datetime.now().strftime("%d-%m-%Y")))
        result=self.cursor.fetchall()
        for data in result:
            excel_sheet.append(data)

        excel_sheet = excel_file.create_sheet(title='Series', index=0)
        excel_sheet.append(self.series_fieldnames)
        query="select Series_id,Season_id,Title,Show_type,Description,Release_year,original_title,Cast,Season_number,Genres,Rating,Age_Rating,Service_name,Added_to_site,Updated_at from {table_name} where Updated_at='%s'".format(table_name=db_detail.Series_table)
        self.cursor.execute(query%(datetime.now().strftime("%d-%m-%Y")))
        result=self.cursor.fetchall()
        for data in result:
            excel_sheet.append(data)

        excel_sheet = excel_file.create_sheet(title='Episodes', index=0)
        excel_sheet.append(self.episode_fileldnames)
        query="select Series_id,Season_id,Episode_id,Series_title,Title,Show_type,Description ,OTT,Duration,season_number,Episode_number,Service_name,Updated_at from {table_name} where Updated_at='%s'".format(table_name=db_detail.Episodes_table)
        self.cursor.execute(query%(datetime.now().strftime("%d-%m-%Y")))
        result=self.cursor.fetchall()
        for data in result:
            excel_sheet.append(data)    

        excel_file.save(os.getcwd()+result_sheet)

        excel_file.close()
        self.connection.close()    



